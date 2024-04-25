import os
import json
import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.remote.webelement import WebElement
from selenium.common.exceptions import *
from time import sleep
from openpyxl import Workbook
from openpyxl import load_workbook

def find_element(driver : webdriver.Chrome, whichBy, unique : str) -> WebElement:
    while True:
        try:
            element = driver.find_element(whichBy, unique)
            break
        except:
            pass
        sleep(1)
    return element

def find_elements(driver : webdriver.Chrome, whichBy, unique : str) -> list[WebElement]:
    while True:
        try:
            elements = driver.find_elements(whichBy, unique)
            break
        except:
            pass
        sleep(0.1)
    return elements

def get_product_url(driver : webdriver.Chrome):

    url = 'https://www.vinci-play.com/en/playground-equipment'
    driver.get(url)
    sleep(1)
    # driver.find_element(By.CSS_SELECTOR, "button[]").click() 
    print("Loaded Vinci")
    # driver.find_element(By.CSS_SELECTOR, '#window2524 > button').click()
    print("Passed the ad window")
    sleep(2)

    # product_addresses = driver.find_elements(By.CSS_SELECTOR, 'li.single-product-item a').get_attribute("href")

    # Find the <a> tag elements
    # elements = driver.find_elements(By.CSS_SELECTOR, 'li.single-product-item a')

    # # Get the href attribute value for each element
    # href_values = [element.get_attribute("href") for element in elements]

    # # Print the href values
    # for href_value in href_values:
    #     # print(href_value)
    #     sheet.append([href_value])

    # workbook.save(os.path.normpath("product-url.xlsx"))
    # print('Saved the Product-url')



workbook = Workbook()
sheet = workbook.active

driver = webdriver.Chrome()
driver.maximize_window()

get_product_url(driver)

wb = load_workbook('product-url.xlsx')
sheet = wb.active

results = []
count = 0

rowCount = sheet.max_row + 1

for row in range(1, rowCount):
    result = {}
    url = sheet.cell(row=row, column=1).value
    # Navigate to the URL
    driver.get(url)
    # Find the <a> tag elements
    product_title = driver.find_element(By.CLASS_NAME, 'product-title').text
    result["title"] = product_title
    trs = driver.find_elements(By.TAG_NAME, 'tr')
    for tr in trs:
        key = tr.text.split('\n')[0]
        val = tr.text.split('\n')[1]
        result[key] = val

    slider_div  = driver.find_element(By.ID, 'slider')
    product_details = slider_div.find_elements(By.CSS_SELECTOR, 'li.slide_thumb_1 img')
    srcs = [product_detail.get_attribute("src") for product_detail in product_details]
    print(srcs)

    # for index, product_detail in enumerate(product_details):
    #     src = product_detail.get_attribute("src")
    #     image_url = src.split("?")[0]  # Remove any query parameters from the URL
    #     folder_name = f"images"  # Set the desired folder name for the image
    #     os.makedirs(folder_name, exist_ok=True)  # Create a new folder for the image
    #     file_name = f"{folder_name}/{image_url.split('/')[-1]}"  # Set the desired file name for the image
    #     response = requests.get(image_url)
    #     with open(file_name, "wb") as file:
    #         file.write(response.content)

    result["url"] = ", ".join(srcs)
    results.append(result)
    count += 1
    print(count)
    with open('results.json', 'w') as file:
        json.dump(results, file)
